# Trabalhando com Arquivos em Python

# Modos de abertura de arquivos: 'r' (leitura), 'w' (escrita), 'a' (append), 'r+' (leitura + escrita).

# Métodos essenciais: open(), read(), write(), close().

# Exemplo de uso:

# Criando e escrevendo em um arquivo
with open('dados.txt', 'w') as arquivo:
    arquivo.write("Nome: João\nIdade: 30\nProfissão: Desenvolvedor")

# Lendo o arquivo
with open('dados.txt', 'r') as arquivo:
    conteudo = arquivo.read()
    print(conteudo)

# Adicionando mais dados (append)
with open('dados.txt', 'a') as arquivo:
    arquivo.write("\nEmail: joao@email.com")

# Excluindo o arquivo (importar módulo 'os')
import os
if os.path.exists('dados.txt'):
    os.remove('dados.txt')
    print("Arquivo removido!")
else:
    print("Arquivo não encontrado.")


###

#Exemplo com funções:

with open('dados.txt', 'r') as arquivo:
    for linha in arquivo:
        print(linha.strip())  #remove '\n'

def ler_arquivo(nome_arquivo):
    """Lê e retorna o conteúdo de um arquivo."""
    with open(nome_arquivo, 'r') as arquivo:
        return arquivo.read()

def escrever_arquivo(nome_arquivo, conteudo):
    """Escreve conteúdo em um arquivo."""
    with open(nome_arquivo, 'w') as arquivo:
        arquivo.write(conteudo)

# Testando as funções
escrever_arquivo('exemplo.txt', 'BOMBARDILLO CROCODILLO')
print(ler_arquivo('exemplo.txt'))

###

# Exemplo para Atividade em Aula:
def contar_linhas(nome_arquivo):
    """Retorna o número de linhas de um arquivo."""
    with open(nome_arquivo, 'r') as arquivo:
        return len(arquivo.readlines())

print(contar_linhas('usuarios.txt'))



import csv

# Escrevendo em CSV
with open('dados.csv', 'w', newline='') as arquivo_csv:
    escritor = csv.writer(arquivo_csv)
    escritor.writerow(["Nome", "Idade", "Cidade"])
    escritor.writerow(["Ana", 25, "Sao Paulo"])
    escritor.writerow(["Carlos", 30, "Rio de Janeiro"])

# Lendo CSV
with open('dados.csv', 'r') as arquivo_csv:
    leitor = csv.reader(arquivo_csv)
    for linha in leitor:
        print(linha)

#EXTRA: Lendo CSV com pandas
import pandas as pd # Para instalar: pip install pandas

df = pd.read_csv('dados.csv')
print(df.head())  # Exibe as primeiras linhas


from openpyxl import Workbook # Para instalar: pip install openpyxl

# Criando uma planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Funcionários"

# Adicionando dados
ws.append(["Nome", "Cargo", "Salário"])
ws.append(["Maria", "Engenheira", 8500])
ws.append(["Pedro", "Analista", 6000])

# Salvando o arquivo
wb.save("funcionarios.xlsx")

# Lendo o arquivo
from openpyxl import load_workbook
wb = load_workbook("funcionarios.xlsx")
ws = wb.active
for linha in ws.iter_rows(values_only=True):
    print(linha)

#EXTRA: lendo xlsx com pandas
import pandas as pd # Para instalar: pip install pandas

df = pd.read_excel('funcionarios.xlsx')
print(df.head())  # Exibe as primeiras linhas